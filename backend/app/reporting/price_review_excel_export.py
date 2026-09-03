"""
Multi-sheet Excel export for a completed price review (spec Section 29). Genuinely testable in
this sandbox - openpyxl is available - unlike the DB/API layer. Operates on plain dataclasses,
not SQLAlchemy models, so it can be exercised standalone (see scripts/demo_price_review.py) and
from the real service layer once that's wired to a DB session.

The one rule this module exists to enforce: "numeric values must remain numeric... do not export
financial amounts as strings" (spec Section 29). Every currency/quantity cell is written as a
Python Decimal/float via openpyxl, never str(value) - checked explicitly in
tests_pure/test_excel_export.py by reading the workbook back and asserting cell.data_type == 'n'.
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ExportLine:
    old_supplier_sku: str | None
    old_description: str | None
    old_pack_raw: str | None
    old_price: Decimal | None
    new_supplier_sku: str | None
    new_description: str | None
    new_pack_raw: str | None
    new_price: Decimal | None
    normalized_old_price: Decimal | None
    normalized_new_price: Decimal | None
    change_amount: Decimal | None
    change_pct: Decimal | None
    historical_volume: Decimal | None
    annual_volume: Decimal | None
    annual_impact: Decimal | None
    margin_impact: Decimal | None
    match_confidence: Decimal | None
    pack_changed: bool
    risk: str | None
    movement_type: str
    buyer_decision: str | None
    target_price: Decimal | None
    potential_cost_avoidance: Decimal | None


@dataclass
class ExportSummary:
    supplier_name: str
    effective_date: str | None
    total_previous_skus: int
    total_new_skus: int
    matched_skus: int
    new_skus: int
    discontinued_skus: int
    increasing_skus: int
    decreasing_skus: int
    unchanged_skus: int
    pack_changes: int
    weighted_average_price_increase_pct: Decimal | None
    annual_cost_impact: Decimal
    products_requiring_manual_review: int


_HEADER_FONT = Font(bold=True)

_MAIN_TABLE_HEADERS = [
    "Supplier SKU (old)", "Supplier SKU (new)", "Product", "Old Pack", "New Pack",
    "Old Price", "New Price", "Normalized Old Price", "Normalized New Price",
    "Change (R)", "Change (%)", "Historical Volume", "Annual Volume",
    "Annual Impact (R)", "Margin Impact", "Match Confidence", "Pack Change",
    "Risk", "Buyer Decision", "Target Price", "Potential Cost Avoidance",
]


def _numeric(value: Decimal | float | None) -> float | None:
    """Coerces to a plain float for openpyxl - Decimal writes fine too, but float is the more
    portable numeric type across Excel versions and avoids any ambiguity about cell.data_type."""
    return float(value) if value is not None else None


def _write_main_row(ws: Worksheet, line: ExportLine) -> None:
    ws.append([
        line.old_supplier_sku, line.new_supplier_sku,
        line.new_description or line.old_description,
        line.old_pack_raw, line.new_pack_raw,
        _numeric(line.old_price), _numeric(line.new_price),
        _numeric(line.normalized_old_price), _numeric(line.normalized_new_price),
        _numeric(line.change_amount), _numeric(line.change_pct),
        _numeric(line.historical_volume), _numeric(line.annual_volume),
        _numeric(line.annual_impact), _numeric(line.margin_impact),
        _numeric(line.match_confidence), ("Yes" if line.pack_changed else "No"),
        line.risk, line.buyer_decision,
        _numeric(line.target_price), _numeric(line.potential_cost_avoidance),
    ])


def _add_table_sheet(wb: openpyxl.Workbook, title: str, lines: list[ExportLine]) -> None:
    ws = wb.create_sheet(title)
    ws.append(_MAIN_TABLE_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for line in lines:
        _write_main_row(ws, line)
    ws.freeze_panes = "A2"


def export_price_review(
    lines: list[ExportLine], summary: ExportSummary, output_path: str | Path
) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # replace the default sheet with named ones in a deliberate order

    # --- Executive Summary ---
    exec_ws = wb.create_sheet("Executive Summary")
    exec_rows = [
        ("Supplier", summary.supplier_name),
        ("Effective Date", summary.effective_date or ""),
        ("Total Previous SKUs", summary.total_previous_skus),
        ("Total New SKUs", summary.total_new_skus),
        ("Matched SKUs", summary.matched_skus),
        ("New Products", summary.new_skus),
        ("Discontinued Products", summary.discontinued_skus),
        ("Increasing", summary.increasing_skus),
        ("Decreasing", summary.decreasing_skus),
        ("Unchanged", summary.unchanged_skus),
        ("Pack Changes", summary.pack_changes),
        ("Weighted Average Price Increase (%)", _numeric(summary.weighted_average_price_increase_pct)),
        ("Annual Cost Impact (R)", _numeric(summary.annual_cost_impact)),
        ("Products Requiring Manual Review", summary.products_requiring_manual_review),
    ]
    for label, value in exec_rows:
        exec_ws.append([label, value])
    exec_ws.column_dimensions["A"].width = 38
    for row in exec_ws.iter_rows(min_col=1, max_col=1):
        row[0].font = _HEADER_FONT

    # --- Full Price Review ---
    _add_table_sheet(wb, "Full Price Review", lines)

    # --- Filtered sheets, per spec Section 29 ---
    _add_table_sheet(wb, "Price Increases", [l for l in lines if l.movement_type == "price_increase"])
    _add_table_sheet(wb, "Price Decreases", [l for l in lines if l.movement_type == "price_decrease"])
    _add_table_sheet(wb, "Pack Changes", [l for l in lines if l.pack_changed])
    _add_table_sheet(wb, "Unmatched Products", [l for l in lines if l.movement_type == "review_required"])
    _add_table_sheet(wb, "New Products", [l for l in lines if l.movement_type == "new_product"])
    _add_table_sheet(wb, "Discontinued Products", [l for l in lines if l.movement_type == "discontinued"])
    _add_table_sheet(wb, "Negotiation Targets", [l for l in lines if l.target_price is not None])

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
