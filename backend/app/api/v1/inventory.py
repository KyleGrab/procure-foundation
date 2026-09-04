"""
Inventory valuation upload route. Real corrections from this codebase's actual conventions,
verified before writing this file, not assumed:
- ValidationFailedError -> HTTP 422 (this codebase's real convention, via main.py's generic
  ProcureIQError handler reading exc.status_code), not 400 - using a non-standard 400 here would
  be the one endpoint in this app that disagrees with every other validation-erroring route.
- Permission.UPLOAD_DATA - no MANAGE_INVENTORY permission exists; this is the real, closest fit.
- "Reject client-supplied organisation overrides" is not separate logic to write: active_org_id
  is baked into the signed JWT itself (app/core/security.py) and is never read from any other
  part of the request - there is no header or body field for it to override in the first place.

.xlsx and .csv only, explicitly not legacy binary .xls - parsing BIFF8/CFBF from scratch was
already declined two turns ago as unverifiable; the operational conversion step happens before a
file reaches this endpoint, not inside it.
"""
from __future__ import annotations

import csv
import io
import uuid
from datetime import date

from fastapi import APIRouter, Depends, File, Form, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import Permission
from app.core.exceptions import NotFoundError, ValidationFailedError
from app.core.permissions import require_permission
from app.core.security import AccessTokenClaims
from app.db.models import Location
from app.db.session import get_db
from app.ingestion.inventory_valuation_mapping import (
    apply_mapping,
    suggest_inventory_valuation_mapping,
)
from app.ingestion.inventory_valuation_validation import validate_inventory_valuation_rows
from app.ingestion.validation import serialize_validation_issues
from app.services.inventory_valuation_service import ingest_inventory_valuation

router = APIRouter(prefix="/inventory", tags=["inventory"])

_REQUIRED_FIELDS = ("supplier_sku", "quantity_on_hand", "unit_cost")


def _parse_upload_to_rows(filename: str, content: bytes) -> list[dict]:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        header = [str(c).strip() if c is not None else "" for c in header_row]
        return [dict(zip(header, row)) for row in rows_iter if any(v is not None for v in row)]
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if lower.endswith(".xls"):
        raise ValidationFailedError(
            "Legacy .xls (binary BIFF8) files are not supported - convert to .xlsx or .csv before uploading",
        )
    raise ValidationFailedError(f"Unsupported file type: {filename!r} - expected .xlsx or .csv")


@router.post("/upload-valuation", status_code=201)
async def upload_inventory_valuation(
    location_id: uuid.UUID = Form(...),
    snapshot_date: date = Form(...),
    is_correction: bool = Form(False),
    file: UploadFile = File(...),
    claims: AccessTokenClaims = Depends(require_permission(Permission.UPLOAD_DATA)),
    db: AsyncSession = Depends(get_db),
) -> dict:
    # Resolve the PUBLIC location_id (what the client sends) to the internal integer id the
    # service layer needs - never trust a client-supplied primary key directly (§1), same
    # pattern as every other route resolving a public_id, scoped to claims.active_org_id so a
    # location belonging to a different organisation can never be targeted even by guessing a
    # valid UUID.
    location_result = await db.execute(
        select(Location.id)
        .where(Location.public_id == location_id)
        .where(Location.organisation_id == claims.active_org_id)
    )
    internal_location_id = location_result.scalar_one_or_none()
    if internal_location_id is None:
        raise NotFoundError(f"Location {location_id} not found")

    content = await file.read()
    raw_rows = _parse_upload_to_rows(file.filename, content)
    if not raw_rows:
        raise ValidationFailedError("Uploaded file contains no data rows")

    mapping = suggest_inventory_valuation_mapping(list(raw_rows[0].keys()))
    unmapped_required = [f for f in _REQUIRED_FIELDS if mapping.get(f) is None]
    if unmapped_required:
        raise ValidationFailedError(
            f"Could not map required columns: {unmapped_required} - rename the source columns "
            f"or extend app.ingestion.inventory_valuation_mapping's alias list",
        )

    mapped_rows = [apply_mapping(row, mapping) for row in raw_rows]
    validation_results = validate_inventory_valuation_rows(mapped_rows)

    all_issues = [issue for result in validation_results for issue in result["issues"]]
    invalid_results = [r for r in validation_results if not r["is_valid"]]
    if invalid_results:
        # The whole batch is rejected, not just the invalid rows silently dropped - a partial
        # ingest would mean the caller never learns some of their real inventory data went
        # missing. Every issue (errors AND warnings) is returned together so one round trip
        # surfaces everything worth fixing, not just the first blocking error.
        raise ValidationFailedError(
            f"{len(invalid_results)} of {len(validation_results)} rows failed validation",
            details=serialize_validation_issues(all_issues),
        )

    validated_records = [r["parsed"] for r in validation_results]

    result = await ingest_inventory_valuation(
        db, organisation_id=claims.active_org_id, user_id=claims.user_id, location_id=internal_location_id,
        snapshot_date=snapshot_date, validated_records=validated_records,
        source_file_storage_key=file.filename, is_correction=is_correction,
    )

    return {
        "record_count": result["record_count"],
        "total_asset_valuation": str(result["total_asset_valuation"]),
        "snapshot_ids": result["snapshot_ids"],
        # Non-blocking warnings (e.g. negative quantity_on_hand) still surfaced even on success -
        # a 201 with silently-swallowed warnings would hide something worth a human's attention.
        "warnings": serialize_validation_issues(all_issues),
    }
