"""Covers spec Section 39 "Excel export" required test, and the explicit Section 29 rule that
financial amounts must stay numeric, never strings."""
from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

import openpyxl

from app.reporting.price_review_excel_export import ExportLine, ExportSummary, export_price_review


def _sample_lines() -> list[ExportLine]:
    return [
        ExportLine(
            old_supplier_sku="SKU1", old_description="Widget A", old_pack_raw="1kg",
            old_price=Decimal("100"), new_supplier_sku="SKU1", new_description="Widget A",
            new_pack_raw="1kg", new_price=Decimal("110"),
            normalized_old_price=Decimal("100"), normalized_new_price=Decimal("110"),
            change_amount=Decimal("10"), change_pct=Decimal("0.10"),
            historical_volume=Decimal("500"), annual_volume=Decimal("500"),
            annual_impact=Decimal("5000"), margin_impact=None,
            match_confidence=Decimal("1.0"), pack_changed=False, risk="medium",
            movement_type="price_increase", buyer_decision="negotiate",
            target_price=Decimal("104"), potential_cost_avoidance=Decimal("3000"),
        ),
        ExportLine(
            old_supplier_sku="SKU2", old_description="Widget B", old_pack_raw="6 x 2L",
            old_price=Decimal("360"), new_supplier_sku="SKU2", new_description="Widget B",
            new_pack_raw="4 x 2L", new_price=Decimal("264"),
            normalized_old_price=Decimal("30"), normalized_new_price=Decimal("33"),
            change_amount=Decimal("3"), change_pct=Decimal("0.10"),
            historical_volume=Decimal("1000"), annual_volume=Decimal("1000"),
            annual_impact=Decimal("3000"), margin_impact=None,
            match_confidence=Decimal("0.98"), pack_changed=True, risk="high",
            movement_type="pack_change", buyer_decision=None,
            target_price=None, potential_cost_avoidance=None,
        ),
        ExportLine(
            old_supplier_sku=None, old_description=None, old_pack_raw=None, old_price=None,
            new_supplier_sku="SKU3", new_description="New Widget C", new_pack_raw="1kg",
            new_price=Decimal("50"), normalized_old_price=None, normalized_new_price=Decimal("50"),
            change_amount=None, change_pct=None, historical_volume=None, annual_volume=None,
            annual_impact=None, margin_impact=None, match_confidence=None, pack_changed=False,
            risk="unclassified", movement_type="new_product", buyer_decision=None,
            target_price=None, potential_cost_avoidance=None,
        ),
    ]


def _sample_summary() -> ExportSummary:
    return ExportSummary(
        supplier_name="Cape Valley Foods (Pty) Ltd", effective_date="2026-09-01",
        total_previous_skus=2, total_new_skus=3, matched_skus=2, new_skus=1, discontinued_skus=0,
        increasing_skus=1, decreasing_skus=0, unchanged_skus=0, pack_changes=1,
        weighted_average_price_increase_pct=Decimal("0.10"), annual_cost_impact=Decimal("8000"),
        products_requiring_manual_review=0,
    )


class TestExcelExport(unittest.TestCase):
    def test_export_creates_all_required_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_price_review(_sample_lines(), _sample_summary(), Path(tmp) / "review.xlsx")
            wb = openpyxl.load_workbook(path)
            expected_sheets = {
                "Executive Summary", "Full Price Review", "Price Increases", "Price Decreases",
                "Pack Changes", "Unmatched Products", "New Products", "Discontinued Products",
                "Negotiation Targets",
            }
            self.assertEqual(set(wb.sheetnames), expected_sheets)

    def test_financial_amounts_are_numeric_not_string(self):
        # The explicit spec Section 29 rule this module exists to enforce.
        with tempfile.TemporaryDirectory() as tmp:
            path = export_price_review(_sample_lines(), _sample_summary(), Path(tmp) / "review.xlsx")
            wb = openpyxl.load_workbook(path)
            ws = wb["Full Price Review"]

            header = [c.value for c in ws[1]]
            price_col = header.index("Old Price") + 1
            impact_col = header.index("Annual Impact (R)") + 1

            first_data_row = ws[2]
            self.assertEqual(first_data_row[price_col - 1].data_type, "n")
            self.assertEqual(first_data_row[impact_col - 1].data_type, "n")
            self.assertIsInstance(first_data_row[price_col - 1].value, (int, float))

    def test_filtered_sheets_contain_only_matching_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_price_review(_sample_lines(), _sample_summary(), Path(tmp) / "review.xlsx")
            wb = openpyxl.load_workbook(path)
            # 1 header row + 1 data row expected on each filtered sheet given the fixture above
            self.assertEqual(wb["Pack Changes"].max_row, 2)
            self.assertEqual(wb["New Products"].max_row, 2)
            self.assertEqual(wb["Price Decreases"].max_row, 1)  # header only, no decreases planted

    def test_executive_summary_has_headline_figures(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = export_price_review(_sample_lines(), _sample_summary(), Path(tmp) / "review.xlsx")
            wb = openpyxl.load_workbook(path)
            ws = wb["Executive Summary"]
            labels = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]
            self.assertIn("Annual Cost Impact (R)", labels)
            self.assertIn("Weighted Average Price Increase (%)", labels)


if __name__ == "__main__":
    unittest.main()
